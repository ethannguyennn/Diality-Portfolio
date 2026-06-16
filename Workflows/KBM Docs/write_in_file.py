import json
import os
import re

import openpyxl

# ── Configuration ─────────────────────────────────────────────────────────────
RECORDS_FOLDER = r'c:\Scripts\KBM\Records\testset'
JSON_PATH      = r'c:\Scripts\KBM\kbmmap.json'
EXCEL_PATH     = r'c:\Scripts\KBM\whitney testing.xlsx'
# ──────────────────────────────────────────────────────────────────────────────


def load_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def normalize(s):
    return s.replace('_', ' ')


def extract_dia_code(filename):
    """Extract DIA identifier from a filename (e.g. 'DIA WF-736J', 'DIA EF-732L.1').
    Handles underscore variants and prefixes like 'Watermarked_'.
    Requires whitespace after 'DIA' to avoid matching 'KH25120-DIA01' style strings."""
    normalized = normalize(filename)
    match = re.search(r'\bDIA\s+([A-Z0-9][A-Z0-9\-\.]*[A-Z0-9]|[A-Z]+)', normalized)
    if match:
        return f"DIA {match.group(1)}"
    return None


def get_id_from_json(data, dia_code):
    """Return the 6/9-digit identifier string for a DIA code.
    Handles both the original plain-string format and the extended dict format."""
    entry = data.get(dia_code)
    if entry is None:
        return None
    if isinstance(entry, str):
        return entry
    return entry.get('id')


def version_score(filename):
    """Return a sortable tuple for the version in a filename. Higher = newer.
    Handles v.X.Y, vX.Y, and bare vX (e.g. v.13.0, v.16, v1.1)."""
    m = re.search(r'\bv\.?(\d+)(?:\.(\d+))?', filename, re.IGNORECASE)
    if m:
        major = int(m.group(1))
        minor = int(m.group(2)) if m.group(2) else 0
        return (major, minor)
    return (0, 0)


def collect_files(folder):
    """Walk folder recursively and return (pdf_files, docx_files) as lists of absolute paths."""
    pdf_files  = []
    docx_files = []
    for root, _, files in os.walk(folder):
        for name in files:
            lower = name.lower()
            full  = os.path.abspath(os.path.join(root, name))
            if lower.endswith('.pdf'):
                pdf_files.append(full)
            elif lower.endswith('.docx'):
                docx_files.append(full)
    return pdf_files, docx_files


def build_docx_index(docx_files):
    """Build a lookup from a leading ID prefix -> list of docx paths.
    The prefix is the leading 6-digit number (optionally followed by -NNN or -NNNN)
    as it appears in the filename."""
    index = {}
    pattern = re.compile(r'^(\d{6}(?:-\d{3,4})?)\b')
    for path in docx_files:
        name = os.path.basename(path)
        m = pattern.match(name)
        if m:
            prefix = m.group(1)
            index.setdefault(prefix, []).append(path)
    return index


def create_workbook():
    """Always create a fresh workbook with headers, overwriting any existing file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Index', 'PDF Path', 'DOCX Path'])
    # 870 px → (870 - 5) / 7 ≈ 123.6 character units
    ws.column_dimensions['B'].width = 123.6
    ws.column_dimensions['C'].width = 123.6
    return wb, ws


def main():
    data       = load_json(JSON_PATH)
    pdf_files, docx_files = collect_files(RECORDS_FOLDER)
    docx_index = build_docx_index(docx_files)
    wb, ws     = create_workbook()
    start_row  = 2
    index      = 1

    no_dia  = 0
    no_json = 0

    # Group PDFs by DIA code, keeping only the highest-versioned file per code
    best_pdf = {}  # {dia_code -> (version_tuple, absolute_path)}
    for pdf_path in pdf_files:
        name     = os.path.basename(pdf_path)
        dia_code = extract_dia_code(name)
        if dia_code is None:
            no_dia += 1
            continue
        score = version_score(name)
        if dia_code not in best_pdf or score > best_pdf[dia_code][0]:
            best_pdf[dia_code] = (score, pdf_path)

    no_docx = 0
    matched = 0

    for dia_code, (score, pdf_path) in sorted(best_pdf.items()):
        doc_id = get_id_from_json(data, dia_code)
        if doc_id is None:
            no_json += 1
            print(f'  [no JSON match]  {os.path.basename(pdf_path)}  →  {dia_code}')
            continue

        candidates = docx_index.get(doc_id, [])
        docx_path  = candidates[0] if candidates else None

        if docx_path is None:
            no_docx += 1

        ws.cell(row=start_row, column=1, value=index)
        ws.cell(row=start_row, column=2, value=pdf_path)
        ws.cell(row=start_row, column=3, value=docx_path if docx_path else '')

        index     += 1
        start_row += 1
        matched   += 1

    wb.save(EXCEL_PATH)

    print(f'\nDone. Wrote {matched} rows to {EXCEL_PATH}')
    print(f'  PDFs without DIA identifier:  {no_dia}')
    print(f'  DIA codes not in JSON:        {no_json}')
    print(f'  DIA codes with no DOCX match: {no_docx}')


if __name__ == '__main__':
    main()
