import os
import re
import shutil

import openpyxl
import pandas as pd

import core.constants as C


def normalize(s):
    return s.replace('_', ' ')


def extract_dia_code(filename):
    normalized = normalize(filename)
    match = re.search(r'\bDIA\s+([A-Z0-9][A-Z0-9\-\.]*[A-Z0-9]|[A-Z]+)', normalized)
    if match:
        return f"DIA {match.group(1)}"
    return None


def get_id_from_json(data, dia_code, normalized_filename=None):
    entry = data.get(dia_code)
    if entry is not None:
        return entry if isinstance(entry, str) else entry.get('id')
    if normalized_filename:
        for key, value in data.items():
            if key in normalized_filename:
                return value if isinstance(value, str) else value.get('id')
    return None


def version_score(filename):
    m = re.search(r'\bv\.?(\d+)(?:\.(\d+))?', filename, re.IGNORECASE)
    if m:
        major = int(m.group(1))
        minor = int(m.group(2)) if m.group(2) else 0
        return (major, minor)
    return (0, 0)


def collect_files(folder):
    pdf_files  = []
    docx_files = []
    skip = {'outdated', 'redline', 'previous revisions'}
    for root, dirs, files in os.walk(folder):
        dirs[:] = [d for d in dirs if d.lower() not in skip]
        for name in files:
            lower = name.lower()
            full  = os.path.abspath(os.path.join(root, name))
            if lower.endswith('.pdf'):
                pdf_files.append(full)
            elif lower.endswith('.docx'):
                docx_files.append(full)
    return pdf_files, docx_files


def build_docx_index(docx_files):
    index   = {}
    pattern = re.compile(r'^(\d{6}(?:-\d{3,4})?)\b')
    for path in docx_files:
        name = os.path.basename(path)
        m    = pattern.match(name)
        if m:
            prefix = m.group(1)
            index.setdefault(prefix, []).append(path)
    return index


def create_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Index', 'PDF Path', 'DOCX Path'])
    ws.column_dimensions['B'].width = 124
    ws.column_dimensions['C'].width = 124
    return wb, ws


def move_outdated_pdfs(all_pdfs_by_id, best_pdf):
    moved = 0
    os.makedirs(C.OUTDATED_FOLDER, exist_ok=True)
    for doc_id, entries in all_pdfs_by_id.items():
        if len(entries) <= 1:
            continue
        best_path = best_pdf[doc_id][1]
        for _, pdf_path in entries:
            if pdf_path == best_path:
                continue
            dest = os.path.join(C.OUTDATED_FOLDER, os.path.basename(pdf_path))
            shutil.move(pdf_path, dest)
            print(f'  [outdated] moved: {os.path.basename(pdf_path)}')
            moved += 1
    return moved


def update_excel(excel_path):
    data = C.KBM_MAP
    pdf_files, docx_files = collect_files(C.RECORDS_FOLDER)

    docx_files = [p for p in docx_files if not re.search(r'red', os.path.basename(p), re.IGNORECASE)]
    docx_index = build_docx_index(docx_files)
    wb, ws     = create_workbook()
    start_row  = 2
    index      = 1

    no_dia  = 0
    no_json = 0

    all_pdfs_by_id = {}
    for pdf_path in pdf_files:
        name     = os.path.basename(pdf_path)
        dia_code = extract_dia_code(name)
        if dia_code is None:
            no_dia += 1
            continue
        doc_id = get_id_from_json(data, dia_code, normalize(name))
        if doc_id is None:
            no_json += 1
            print(f'  [no JSON match]  {name}  ->  {dia_code}')
            continue
        score = version_score(name)
        all_pdfs_by_id.setdefault(doc_id, []).append((score, pdf_path))

    best_pdf = {
        doc_id: max(entries, key=lambda x: x[0])
        for doc_id, entries in all_pdfs_by_id.items()
    }

    moved_count = move_outdated_pdfs(all_pdfs_by_id, best_pdf)

    no_docx         = 0
    matched         = 0
    missing_doc_ids = []

    for doc_id, (score, pdf_path) in sorted(best_pdf.items()):
        candidates = docx_index.get(doc_id, [])
        docx_path  = candidates[0] if candidates else None

        if docx_path is None:
            no_docx += 1
            missing_doc_ids.append((doc_id, os.path.basename(pdf_path)))

        ws.cell(row=start_row, column=1, value=index)
        ws.cell(row=start_row, column=2, value=pdf_path)
        ws.cell(row=start_row, column=3, value=docx_path if docx_path else '')

        index     += 1
        start_row += 1
        matched   += 1

    wb.save(excel_path)

    print(f'\nDone. Wrote {matched} rows to {excel_path}')
    print(f'  Outdated PDFs moved to outdated/: {moved_count}')
    print(f'  PDFs without DIA identifier:       {no_dia}')
    print(f'  DIA codes not in JSON:             {no_json}')
    print(f'  DIA codes with no DOCX match:      {no_docx}')

    if missing_doc_ids:
        print('\nDOCX files needed (add to folder):')
        for doc_id, pdf_name in missing_doc_ids:
            print(f'  {doc_id} -- {pdf_name}')

    return missing_doc_ids
