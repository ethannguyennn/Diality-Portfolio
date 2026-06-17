import json
import os
import re
import shutil

import openpyxl

# ── Configuration ─────────────────────────────────────────────────────────────
RECORDS_FOLDER  = r'c:\Scripts\KBM\Records\testset'
OUTDATED_FOLDER = os.path.join(RECORDS_FOLDER, 'outdated')
JSON_PATH       = r'c:\Scripts\KBM\kbmmap.json'
EXCEL_PATH      = r'c:\Scripts\KBM\whitney testing.xlsx'
# ──────────────────────────────────────────────────────────────────────────────


def load_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def normalize(s):
    return s.replace('_', ' ')


def extract_dia_code(filename):
    # Extract DIA identifier from a filename 
    normalized = normalize(filename)
    match = re.search(r'\bDIA\s+([A-Z0-9][A-Z0-9\-\.]*[A-Z0-9]|[A-Z]+)', normalized)
    if match:
        return f"DIA {match.group(1)}"
    return None


def get_id_from_json(data, dia_code):
    entry = data.get(dia_code)
    if entry is None:
        return None
    if isinstance(entry, str):
        return entry
    return entry.get('id')


def version_score(filename):
    # Return a sortable tuple for the version in a filename. 
    m = re.search(r'\bv\.?(\d+)(?:\.(\d+))?', filename, re.IGNORECASE)
    if m:
        major = int(m.group(1))
        minor = int(m.group(2)) if m.group(2) else 0
        return (major, minor)
    return (0, 0)


def collect_files(folder):
    pdf_files  = []
    docx_files = []
    skip = {'outdated', 'redline'}
    for root, dirs, files in os.walk(folder):
        dirs[:] = [d for d in dirs if d.lower() not in skip]
        for name in files:
            lower = name.lower()
            full  = os.path.abspath(os.path.join(root, name))
            if lower.endswith('.pdf'):
                pdf_files.append(full)
            elif lower.endswith('.docx'):
                docx_files.append(full)
    return pdf_files, docx_files


def build_docx_index(docx_files):
    #Build a lookup from a leading ID prefix -> list of docx paths.
    index = {}
    pattern = re.compile(r'^(\d{6}(?:-\d{3,4})?)\b')
    for path in docx_files:
        name = os.path.basename(path)
        m = pattern.match(name)
        if m:
            prefix = m.group(1)
            index.setdefault(prefix, []).append(path)
    return index


def create_workbook():
    # Always create a fresh workbook with headers, overwriting any existing file.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Index', 'PDF Path', 'DOCX Path'])
    ws.column_dimensions['B'].width = 124
    ws.column_dimensions['C'].width = 124
    return wb, ws



def move_outdated_pdfs(all_pdfs_by_dia, best_pdf):
    # Move every non-best PDF version into OUTDATED_FOLDER, preserving subfolder structure.
    moved = 0
    os.makedirs(OUTDATED_FOLDER, exist_ok=True)
    for dia_code, entries in all_pdfs_by_dia.items():
        if len(entries) <= 1:
            continue
        best_path = best_pdf[dia_code][1]
        for _, pdf_path in entries:
            if pdf_path == best_path:
                continue
            dest = os.path.join(OUTDATED_FOLDER, os.path.basename(pdf_path))
            shutil.move(pdf_path, dest)
            print(f'  [outdated] moved: {os.path.basename(pdf_path)}')
            moved += 1
    return moved


def main():
    data       = load_json(JSON_PATH)
    pdf_files, docx_files = collect_files(RECORDS_FOLDER)

    docx_files = [p for p in docx_files if not re.search(r'red', os.path.basename(p), re.IGNORECASE)]
    docx_index = build_docx_index(docx_files)
    wb, ws     = create_workbook()
    start_row  = 2
    index      = 1

    no_dia  = 0
    no_json = 0

    # Group all PDFs by DIA code
    all_pdfs_by_dia = {}  # {dia_code -> [(version_tuple, absolute_path), ...]}
    for pdf_path in pdf_files:
        name     = os.path.basename(pdf_path)
        dia_code = extract_dia_code(name)
        if dia_code is None:
            no_dia += 1
            continue
        score = version_score(name)
        all_pdfs_by_dia.setdefault(dia_code, []).append((score, pdf_path))

    # Best version per DIA code
    best_pdf = {
        dia_code: max(entries, key=lambda x: x[0])
        for dia_code, entries in all_pdfs_by_dia.items()
    }

    # Move outdated versions
    moved_count = move_outdated_pdfs(all_pdfs_by_dia, best_pdf)

    no_docx        = 0
    matched        = 0
    missing_doc_ids = []

    for dia_code, (score, pdf_path) in sorted(best_pdf.items()):
        doc_id = get_id_from_json(data, dia_code)
        if doc_id is None:
            no_json += 1
            print(f'  [no JSON match]  {os.path.basename(pdf_path)}  →  {dia_code}')
            continue

        candidates = docx_index.get(doc_id, [])
        docx_path  = candidates[0] if candidates else None

        if docx_path is None:
            no_docx += 1
            missing_doc_ids.append(doc_id)

        ws.cell(row=start_row, column=1, value=index)
        ws.cell(row=start_row, column=2, value=pdf_path)
        ws.cell(row=start_row, column=3, value=docx_path if docx_path else '')

        index     += 1
        start_row += 1
        matched   += 1

    wb.save(EXCEL_PATH)

    print(f'\nDone. Wrote {matched} rows to {EXCEL_PATH}')
    print(f'  Outdated PDFs moved to outdated/: {moved_count}')
    print(f'  PDFs without DIA identifier:       {no_dia}')
    print(f'  DIA codes not in JSON:             {no_json}')
    print(f'  DIA codes with no DOCX match:      {no_docx}')

    if missing_doc_ids:
        print('\nDOCX files needed (add to folder):')
        for doc_id in missing_doc_ids:
            print(f'  {doc_id}')


if __name__ == '__main__':
    main()
