import argparse
import copy
import io
import os
import re
import uuid
import shutil
import zipfile
import requests
from requests.auth import HTTPBasicAuth
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Protection

from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList
from datetime import datetime

load_dotenv()

BASE_URL = "https://diality.atlassian.net"
PROJECT = "LDT"
SHEET_NAME = "Sprint Template"
DATA_START_ROW = 4  

SWVV_TEAM = ["Raghu Kallala", "Thomas Lippold", "Tejaskumar Patel", "Tiffany Mejia",
             "Zoltan Miskolci", "Sarina Cheung", "Tisha Patel", "Ethan Nguyen"]
FW_TEAM   = ["Arpita Srivastava", "Jashwant Gantyada", "Michael Garthwaite", "Praneeth Bunne",
             "Sameer Poyil", "Varshini Nagabooshanam", "Vijay Pamula", "Suresh Dharnala", "Santhos Kumar Reddy",
             "Vinayakam Mani", "Sean Nash"]
SW_TEAM   = ["Nicholas Ramirez", "Stephen Quong", "Dara Navaei", "Behrouz NematiPour"]
SYS_TEAM  = ["Eliza Petersen", "Caitlynn Chang", "Christina Heine", "Abhijit Barman", "Chris Yu", "Vitas Buenaventura", "Emiline Hernandez"]
TEAMS = [FW_TEAM, SWVV_TEAM, SW_TEAM, SYS_TEAM]
TEAM_NAMES = ["FW", "SWVV", "SW", "SYS"]
_SYS_TEAM_IDX       = len(TEAMS) - 1  
_UNMATCHED_TEAM_IDX = len(TEAMS)      

NICKNAME_MAP = {
    "Tejaskumar Patel":       "Tejas",
    "Nicholas Ramirez":       "Nico",
    "Thomas Lippold":         "Tom",
    "Jashwant Gantyada":      "Jashwant",       
    "Behrouz NematiPour":     "Behrouz",
    "Sean Nash":              "Sean",
    "Vinayakam Mani":         "Vinay",
    "Raghu Kallala":          "Raghu",
    "Tiffany Mejia":          "Tiffany",
    "Zoltan Miskolci":        "Zoltan",
    "Sarina Cheung":          "Sarina",
    "Tisha Patel":            "Tisha",
    "Ethan Nguyen":           "Ethan",
    "Arpita Srivastava":      "Arpita",
    "Michael Garthwaite":     "Michael",
    "Praneeth Bunne":         "Praneeth",
    "Sameer Poyil":           "Sameer",
    "Varshini Nagabooshanam": "Varshini",
    "Vijay Pamula":           "Vijay",
    "Suresh Dharnala":        "Suresh",
    "Dara Navaei":            "Dara",
    "Eliza Petersen":         "Eliza",
    "Stephen Quong":          "Stephen",
    "Christina Heine":        "Christina",
    "Caitlynn Chang":         "Caitlynn",
    "Santhos Kumar Reddy":    "Santhos",
    "Abhijit Barman":         "Abhijit",
    "Chris Yu":               "Chris",
    "Vitas Buenaventura":     "Vitas",
    "Emiline Hernandez":      "Emiline",
}

_ASSIGNEE_TO_TEAM_IDX: dict = {}
for _ti, _team in enumerate(TEAMS):
    for _full in _team:
        _ASSIGNEE_TO_TEAM_IDX[_full] = _ti
        _nick = NICKNAME_MAP.get(_full)
        if _nick:
            _ASSIGNEE_TO_TEAM_IDX[_nick] = _ti

SHAREPOINT_SYNC_PATH = os.getenv("SHAREPOINT_SYNC_PATH", "").strip()
ARCHIVE_PATH         = os.getenv("ARCHIVE_PATH", "").strip()

# Column headers written to row 3 
SHEET_HEADERS = ["Sprint", "Team", "Original Assignee", "Current Assignee", " Type", "Scope Changes", "Description", "Jira #", " Status"]

TABLE_DISPLAY_NAME = "Table1"
TABLE_STYLE        = "TableStyleMedium7"

TYPE_FONT_COLORS = {
    "Bug":          "FF0000",
    "Dialin ticket": "4472C4",
    "Support":      "00B050",
    "Little-V":     "7030A0",
    "Big-V":        "000000",
}

# Issue types whose Status column uses the Bug/Dialin-specific dropdown
BUG_DIALIN_TYPES = {"Bug", "Dialin ticket"}
BUG_STATUS_MAP = {
    "More Info Needed":                       "More Info Needed",
    "On Hold":                                 "On Hold",
    "Blocked":                                 "Blocked",
    "Ready for SW/FW Implementation":          "Ready for SW/FW Implementation",
    "Investigation In Progress (SW/FW)":       "Investigation in Progress",
    "Fix In Progress":                         "Fix in Progress",
    "Ready for Fix to be Tested":              "Ready for Fix to be Tested",
    "Testing Fix in Progress":                 "Testing Fix in Progress",
    "Test Passed - Ready for Updated Release": "Ready for Updated Release",
    "Test Failed - SW/FW Fix Needed":          "SW/FW Fix Needed",
    "Test Passed - Needs Final Fix Details":   "Needs Final Fix Details",
    "Done":                                    "Done",
}

MIN_SPRINT = 28  

SP_SYNC_COLS = (10, 12)  
JIRA_NOTE_RE  = re.compile(r'xlsx<([^>]*)>xlsx')

VALIDATION_ROW_END    = 1000   
_SELECTION_SPRINT_MAX = 100   

_COL_VALIDATIONS = {
    "A": "Selection!$F$2:$F$33",
    "B": "Selection!$B$2:$B$5",   # Team (new)
    "C": "Selection!$A$2:$A$27",  # Original Assignee (was B)
    "D": "Selection!$A$2:$A$27",  # Current Assignee  (was C)
    "E": "Selection!$D$2:$D$6",   # Type              (was D)
    "F": "Selection!$H$2:$H$4",   # Scope             (was E)
    "G": "Selection!$K$2:$K$3",   # Description Y/N   (was F)
    "J": "Selection!$J$2:$J$5",   # was I
}

def _fetch_paginated(url, jql, auth):
    all_issues = []
    start_at = 0
    while True:
        params = {
            "fields": "summary,status,assignee,issuetype,customfield_10020,description,comment",
            "expand": "changelog",
            "jql": jql,
            "maxResults": 100,
            "startAt": start_at,
        }
        response = requests.get(url, params=params, auth=auth)
        response.raise_for_status()
        data = response.json()
        issues = data["issues"]
        all_issues.extend(issues)
        total = data.get("total", 0)
        start_at += len(issues)
        if start_at >= total or not issues:
            break
    return all_issues

def _sprint_names(s):
    return {n.strip() for n in s.split(",") if n.strip()} if s else set()

def _has_sprint_history(issue):
    # Return True if the changelog contains a Sprint transition into sprint >= MIN_SPRINT.
    for history in issue["changelog"]["histories"]:
        for item in history["items"]:
            if item["field"] != "Sprint":
                continue
            for name in _sprint_names(item.get("toString") or ""):
                if _sprint_num(name) >= MIN_SPRINT:
                    return True
    return False


def _assignee_nickname(display_name: str) -> str:
    # Convert a full Jira display name to the Selection-sheet nickname.
    if not display_name:
        return "Unassigned"
    nick = NICKNAME_MAP.get(display_name)
    if nick:
        return nick
    return display_name.split()[0]

def _team_for_nickname(nickname):
    idx = _ASSIGNEE_TO_TEAM_IDX.get(nickname)
    if idx is None:
        return None, _UNMATCHED_TEAM_IDX
    return TEAM_NAMES[idx], idx


def _adf_to_text(node) -> str:
    if isinstance(node, str):
        return node
    if not isinstance(node, dict):
        return ""
    if node.get("type") == "text":
        return node.get("text", "")
    return "".join(_adf_to_text(c) for c in node.get("content", []))


def _latest_jira_note(issue, auth) -> str:
    comment_field = issue["fields"].get("comment") or {}
    comments = list(comment_field.get("comments", []))
    total = comment_field.get("total", len(comments))

    if total > len(comments):
        url = f"{BASE_URL}/rest/api/2/issue/{issue['key']}/comment"
        resp = requests.get(url, params={"maxResults": 1000}, auth=auth)
        resp.raise_for_status()
        comments = resp.json().get("comments", [])

    best_text = None
    best_created = ""
    for comment in comments:
        body_text = _adf_to_text(comment.get("body") or "")
        m = JIRA_NOTE_RE.search(body_text)
        if m and comment.get("created", "") >= best_created:
            best_created = comment["created"]
            best_text = m.group(1).strip()

    return best_text


def get_jira_issues():
    auth = HTTPBasicAuth(os.getenv("JIRA_EMAIL"), os.getenv("JIRA_API_TOKEN"))

    board_issues = _fetch_paginated(
        f"{BASE_URL}/rest/agile/1.0/board/84/issue",
        "sprint in openSprints() OR sprint in closedSprints() OR sprint in futureSprints()",
        auth,
    )

    seen = {i["key"] for i in board_issues}

    # Fetch every backlog issue and keep only those with sprint changelog entries.
    all_backlog = _fetch_paginated(
        f"{BASE_URL}/rest/agile/1.0/board/84/backlog",
        f"project = {PROJECT}",
        auth,
    )
    backlog_with_history = [
        i for i in all_backlog
        if i["key"] not in seen and _has_sprint_history(i)
    ]

    all_issues = board_issues + backlog_with_history
    for issue in all_issues:
        issue["_jira_note"] = _latest_jira_note(issue, auth)
    return all_issues

def _sprint_num(name):
    m = re.search(r'(\d+)', name)
    return int(m.group(1)) if m else -1

def get_type(issue):
    issuetype = issue["fields"]["issuetype"]["name"].lower()
    summary   = issue["fields"].get("summary", "").lower()

    if "dialin" in issuetype or "dialin" in summary:
        return "Dialin ticket"
    elif issuetype == "bug":
        return "Bug"
    elif "little" in issuetype:
        return "Little-V"
    elif "support" in summary:
        return "Support"
    else:
        return "Big-V"

def parse_date(date_str):
    if not date_str:
        return None
    date_str = date_str.replace("Z", "+00:00")
    date_str = re.sub(r'([+-])(\d{2})(\d{2})$', r'\1\2:\3', date_str)
    return datetime.fromisoformat(date_str)

def _classify_sprint_timing(sprint, histories):
    sprint_name     = sprint["name"]
    sprint_start_dt = parse_date(sprint.get("startDate"))

    if not sprint_start_dt:
        return "Scoped"

    most_recent_add = None
    for history in histories:
        for item in history["items"]:
            if item["field"] != "Sprint":
                continue
            to_sprints   = _sprint_names(item.get("toString")  or "")
            from_sprints = _sprint_names(item.get("fromString") or "")
            if sprint_name in to_sprints and sprint_name not in from_sprints:
                added_dt = parse_date(history["created"])
                if added_dt and (most_recent_add is None or added_dt > most_recent_add):
                    most_recent_add = added_dt

    if most_recent_add is None:
        return "Scoped"

    sprint_start_eod = sprint_start_dt.replace(hour=23, minute=59, second=59, microsecond=0)
    return "Added" if most_recent_add > sprint_start_eod else "Scoped"


def _determine_current_sprint_num(issues, ws) -> int:
    active, closed = [], []
    for issue in issues:
        for s in (issue["fields"].get("customfield_10020") or []):
            num   = _sprint_num(s.get("name", ""))
            state = s.get("state", "")
            if num < MIN_SPRINT:
                continue
            if state == "active":
                active.append(num)
            elif state == "closed":
                closed.append(num)
    if active:
        return max(active)
    try:
        b2 = int(ws.cell(row=2, column=2).value)
        if b2 >= MIN_SPRINT:
            return b2
    except (TypeError, ValueError):
        pass
    if closed:
        return max(closed)
    return MIN_SPRINT


def get_sprint_history(issue, current_sprint_num):
    fields      = issue["fields"]
    histories   = issue["changelog"]["histories"]
    sprint_info = fields.get("customfield_10020") or []

    current_sprint_names = {s["name"] for s in sprint_info}
    current_sprint_objs  = {s["name"]: s for s in sprint_info}

    # Collect every sprint name ever referenced in the changelog
    all_sprint_names = set(current_sprint_names)
    for history in histories:
        for item in history["items"]:
            if item["field"] != "Sprint":
                continue
            for name in _sprint_names(item.get("fromString") or ""):
                all_sprint_names.add(name)
            for name in _sprint_names(item.get("toString") or ""):
                all_sprint_names.add(name)

    sprint_entries = {}  
    for name in all_sprint_names:
        if name not in current_sprint_names:
            num = _sprint_num(name)
            if num >= MIN_SPRINT and num <= current_sprint_num:
                sprint_entries[num] = "Move out of Sprint"

    
    for name, sprint_obj in current_sprint_objs.items():
        num = _sprint_num(name)
        if num >= MIN_SPRINT:
            sprint_entries[num] = _classify_sprint_timing(sprint_obj, histories)

    return sorted(sprint_entries.items(), key=lambda x: x[0])  


def get_status(issue, issue_type) -> str:
    status      = issue["fields"]["status"]
    status_name = status["name"]

    if issue_type in BUG_DIALIN_TYPES:
        return BUG_STATUS_MAP.get(status_name, status_name)

    status_name  = status_name.lower()
    category_key = status["statusCategory"]["key"]

    if "blocked" in status_name:
        return "Blocked"
    if "on hold" in status_name or "on-hold" in status_name:
        return "On Hold"
    if status_name == "n/a":
        return "N/A"
    if category_key == "done" or "done" in status_name or "complete" in status_name:
        return "Done"
    if category_key == "new":
        return "Not started"
    return "In-progress"


def _extract_key(label):
    """Extract Jira key from a cell value like '[LDT-123] Summary title'."""
    m = re.match(r'\[([A-Z]+-\d+)\]', str(label) if label is not None else '')
    return m.group(1) if m else None

def _ensure_sheet_setup(ws):
    for col, header in enumerate(SHEET_HEADERS, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value     = header
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[3].height = 18.75
    ws.freeze_panes = "A4"

    widths = {"A": 9.29, "B": 9, "C": 10.57, "D": 10.57, "E": 11.57, "F": 18.43, "H": 103.14, "I": 16.29}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _ensure_selection_sprints(wb):
    if "Selection" not in wb.sheetnames:
        return
    ws_sel = wb["Selection"]
    for sprint_num in range(1, _SELECTION_SPRINT_MAX + 1):
        row = sprint_num + 1  # row 1 is header; sprint N lives at row N+1
        if ws_sel.cell(row=row, column=6).value is None:
            ws_sel.cell(row=row, column=6).value = sprint_num


def _ensure_data_validations(ws):
    ws.data_validations = DataValidationList()
    for col_letter, formula in _COL_VALIDATIONS.items():
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid entry",
            error="Please select from the dropdown list.",
            errorStyle="warning",
        )
        dv.add(f"{col_letter}{DATA_START_ROW}:{col_letter}{VALIDATION_ROW_END}")
        ws.add_data_validation(dv)


def _ensure_status_validation(ws) -> None:
    # Apply data validation to column I 
    _L_RANGE = "Selection!$L$2:$L$13"
    _E_RANGE = "Selection!$E$2:$E$7"

    row_types = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        e_val = ws.cell(row=row_idx, column=5).value
        row_types.append((row_idx, e_val in BUG_DIALIN_TYPES))

    groups = []
    if row_types:
        cur_start, cur_type = row_types[0]
        for row_idx, is_bug in row_types[1:]:
            if is_bug != cur_type:
                groups.append((cur_start, row_idx - 1, cur_type))
                cur_start, cur_type = row_idx, is_bug
        groups.append((cur_start, row_types[-1][0], cur_type))
        # All rows beyond the last written row get the default E range
        last_data_row = row_types[-1][0]
        if last_data_row < VALIDATION_ROW_END:
            groups.append((last_data_row + 1, VALIDATION_ROW_END, False))
    else:
        groups.append((DATA_START_ROW, VALIDATION_ROW_END, False))

    for start_row, end_row, is_bug_dialin in groups:
        dv = DataValidation(
            type="list",
            formula1=_L_RANGE if is_bug_dialin else _E_RANGE,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid entry",
            error="Please select from the dropdown list.",
            errorStyle="warning",
        )
        dv.add(f"I{start_row}:I{end_row}")
        ws.add_data_validation(dv)


def _original_assignee_for_sprint(issue, sprint_obj) -> str:
    """
    Return the nickname of the assignee at the effective cutoff for this sprint:
      - Scoped tickets (added <= EOD of sprint start): assignee at EOD of the sprint start day.
      - Added tickets (added > EOD of sprint start): assignee at the moment of sprint entry.
    Falls back to current assignee if sprint_obj is None or no sprint entry event is found.
    """
    fields    = issue["fields"]
    histories = issue["changelog"]["histories"]
    current_full = fields["assignee"]["displayName"] if fields.get("assignee") else ""

    if not sprint_obj:
        return _assignee_nickname(current_full)

    sprint_name = sprint_obj.get("name", "")

    sprint_added_dt = None
    for history in histories:
        for item in history["items"]:
            if item["field"] != "Sprint":
                continue
            to_sprints   = _sprint_names(item.get("toString")  or "")
            from_sprints = _sprint_names(item.get("fromString") or "")
            if sprint_name in to_sprints and sprint_name not in from_sprints:
                dt = parse_date(history["created"])
                if sprint_added_dt is None or dt < sprint_added_dt:
                    sprint_added_dt = dt

    if sprint_added_dt is None:
        return _assignee_nickname(current_full)
    sprint_start_dt = parse_date(sprint_obj.get("startDate"))
    if sprint_start_dt:
        sprint_start_eod = sprint_start_dt.replace(hour=23, minute=59, second=59, microsecond=0)
        cutoff_dt = max(sprint_added_dt, sprint_start_eod)
    else:
        cutoff_dt = sprint_added_dt

    assignee_changes = []
    for history in histories:
        for item in history["items"]:
            if item["field"] == "assignee":
                assignee_changes.append((parse_date(history["created"]), item))
    assignee_changes.sort(key=lambda x: x[0])

    if not assignee_changes:
        return _assignee_nickname(current_full)

    current_state = assignee_changes[0][1].get("fromString") or ""
    for change_dt, item in assignee_changes:
        if change_dt <= cutoff_dt:
            current_state = item.get("toString") or ""
        else:
            break

    return _assignee_nickname(current_state) if current_state else "Unassigned"


def _write_issue_row(ws, row_idx, row_record):
    issue      = row_record["issue"]
    sprint_num = row_record["sprint_num"]
    scope      = row_record["scope"]
    team       = row_record.get("team")

    fields = issue["fields"]
    key    = issue["key"]

    full_name         = fields["assignee"]["displayName"] if fields.get("assignee") else ""
    current_assignee  = _assignee_nickname(full_name)
    original_assignee = row_record.get("original_assignee") or current_assignee
    issue_type   = get_type(issue)
    ticket_url   = f"{BASE_URL}/browse/{key}"
    ticket_label = f"[{key}] {fields.get('summary', '')}"
    desc_text       = _adf_to_text(fields.get("description") or "")
    has_description = "Y" if len(desc_text) > 20 else "N"
    status = get_status(issue, issue_type)

    # Column order: A          B     C                  D                E           F      G              H             I
    values = [sprint_num, team, original_assignee, current_assignee, issue_type, scope, has_description, ticket_label, status]

    align = Alignment(horizontal="left", vertical="center")

    for col, value in enumerate(values, start=1):
        cell            = ws.cell(row=row_idx, column=col, value=value)
        cell.alignment  = align
        cell.protection = Protection(locked=False)
        if col == 8:  # H — Jira # hyperlink
            cell.font      = Font(name="Calibri", size=11, color="1155CC", underline="single")
            cell.hyperlink = ticket_url
        elif col == 5:  # E — Type
            cell.font      = Font(name="Calibri", size=11, color=TYPE_FONT_COLORS.get(issue_type, "000000"))
            cell.hyperlink = None
        else:
            cell.hyperlink = None

    jira_note = row_record.get("jira_note")
    if jira_note is not None:
        cell_k            = ws.cell(row=row_idx, column=11, value=jira_note)
        cell_k.alignment  = align
        cell_k.protection = Protection(locked=False)


def _ensure_table_in_model(ws, new_ref: str) -> None:
    """Update the existing table ref, or create a minimal Table if none exists."""
    if ws.tables:
        for tbl in ws.tables.values():
            tbl.ref = new_ref
    else:
        tbl = Table(displayName=TABLE_DISPLAY_NAME, ref=new_ref)
        tbl.tableStyleInfo = TableStyleInfo(name=TABLE_STYLE, showRowStripes=True)
        min_col, _, max_col, _ = range_boundaries(new_ref)
        tbl.tableColumns = [
            TableColumn(
                id=i,
                name=str(cell_val) if (cell_val := ws.cell(row=3, column=col).value) and str(cell_val).strip()
                     else f"Column{col}"
            )
            for i, col in enumerate(range(min_col, max_col + 1), start=1)
        ]
        ws.add_table(tbl)



def _xml_attr(value: str) -> str:
    return (value.replace("&", "&amp;").replace("<", "&lt;")
                 .replace(">", "&gt;").replace('"', "&quot;"))


def _generate_table_xml(ref: str, col_names: list, uid: str = None) -> bytes:
    xr_uid = uid or ("{" + str(uuid.uuid4()).upper() + "}")
    cols_xml = "".join(
        f'<tableColumn id="{i + 1}" name="{_xml_attr(n)}"/>'
        for i, n in enumerate(col_names)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<table xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        ' xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"'
        ' xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision"'
        ' xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"'
        ' mc:Ignorable="xr xr3"'
        f' id="1" name="{TABLE_DISPLAY_NAME}" displayName="{TABLE_DISPLAY_NAME}"'
        f' ref="{ref}" totalsRowShown="0" xr:uid="{xr_uid}">'
        f'<autoFilter ref="{ref}"/>'
        f'<tableColumns count="{len(col_names)}">{cols_xml}</tableColumns>'
        f'<tableStyleInfo name="{TABLE_STYLE}" showFirstColumn="0" showLastColumn="0"'
        ' showRowStripes="1" showColumnStripes="0"/>'
        '</table>'
    ).encode("utf-8")


def _capture_table_xml(xlsx_path: str) -> tuple:
    with zipfile.ZipFile(xlsx_path, "r") as z:
        table_paths = [n for n in z.namelist()
                       if n.startswith("xl/tables/") and n.endswith(".xml")]
        if not table_paths:
            return None, None
        table_path = table_paths[0]
        return table_path, z.read(table_path).decode("utf-8", errors="replace")


def _patch_xlsx_table(xlsx_path: str, final_row: int,
                      orig_table_xml: str = None, col_names: list = None) -> None:
    
    ref = f"A3:L{final_row}"

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        table_paths = [n for n in zin.namelist()
                       if n.startswith("xl/tables/") and n.endswith(".xml")]
        if not table_paths:
            return  # openpyxl wrote no table; nothing to patch

        table_path = table_paths[0]
        raw_xml    = zin.read(table_path).decode("utf-8", errors="replace")
        entries    = [(info, zin.read(info.filename)) for info in zin.infolist()]

    source = orig_table_xml or raw_xml
    uid_m  = re.search(r'xr:uid="([^"]*)"', source)
    clean_xml = _generate_table_xml(ref, col_names or [], uid=uid_m.group(1) if uid_m else None)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for info, data in entries:
            zout.writestr(info.filename, clean_xml if info.filename == table_path else data)

    with open(xlsx_path, "wb") as f:
        f.write(buf.getvalue())

def _insert_point_for_sprint(ws, sprint_num):
    result = DATA_START_ROW - 1
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        try:
            n = int(ws.cell(row=row_idx, column=1).value)
        except (TypeError, ValueError):
            continue
        if n <= sprint_num:
            result = row_idx
    return result


def _sprint_started(sprint_num: int, sprint_map: dict) -> bool:
    # Return True if the sprint has started (active or closed), False if future/unstarted
    s = sprint_map.get(sprint_num)
    if not s:
        return True  
    state = s.get("state", "")
    if state in ("active", "closed"):
        return True
    if state == "future":
        return False
    # Unknown state — fall back to startDate comparison
    start_dt = parse_date(s.get("startDate"))
    if not start_dt:
        return False
    return datetime.now(start_dt.tzinfo) >= start_dt


_SORT_NUM_COLS = 12  


def _snapshot_row(ws, row_idx, num_cols=_SORT_NUM_COLS):
    """Capture full-fidelity per-cell state (value, style, hyperlink, comment) for a row."""
    cells = []
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cells.append({
            "value":         cell.value,
            "font":          copy.copy(cell.font),
            "fill":          copy.copy(cell.fill),
            "border":        copy.copy(cell.border),
            "alignment":     copy.copy(cell.alignment),
            "number_format": cell.number_format,
            "protection":    copy.copy(cell.protection),
            "hyperlink":     copy.copy(cell.hyperlink),
            "comment":       copy.copy(cell.comment),
        })
    return cells


def _restore_row(ws, row_idx, snapshot):
    """Write a snapshot captured by _snapshot_row back into row_idx."""
    for col, state in enumerate(snapshot, start=1):
        cell               = ws.cell(row=row_idx, column=col)
        cell.value         = state["value"]
        cell.font          = state["font"]
        cell.fill          = state["fill"]
        cell.border        = state["border"]
        cell.alignment     = state["alignment"]
        cell.number_format = state["number_format"]
        cell.protection    = state["protection"]
        cell.hyperlink     = state["hyperlink"]
        cell.comment       = state["comment"]


def _sort_sprint_blocks(ws, current_sprint_num) -> None:
    # Find contiguous blocks of rows sharing the same sprint number in column A.
    blocks = []  # (start_row, end_row, sprint_num)
    row_idx = DATA_START_ROW
    while row_idx <= ws.max_row:
        try:
            sprint_num = int(ws.cell(row=row_idx, column=1).value)
        except (TypeError, ValueError):
            row_idx += 1
            continue
        end_row = row_idx
        while end_row + 1 <= ws.max_row:
            try:
                next_sprint = int(ws.cell(row=end_row + 1, column=1).value)
            except (TypeError, ValueError):
                break
            if next_sprint != sprint_num:
                break
            end_row += 1
        blocks.append((row_idx, end_row, sprint_num))
        row_idx = end_row + 1

    def _sort_key(snapshot):
        nickname = snapshot[2]["value"]  # column C — Original Assignee
        _, team_idx = _team_for_nickname(nickname)
        return (team_idx, nickname or "")

    for start_row, end_row, sprint_num in blocks:
        if sprint_num != current_sprint_num:
            continue

        sortable_positions = []
        snapshots = []
        for r in range(start_row, end_row + 1):
            if ws.cell(row=r, column=6).value == "Move out of Sprint":
                continue
            sortable_positions.append(r)
            snapshots.append(_snapshot_row(ws, r))

        if len(sortable_positions) <= 1:
            continue

        snapshots.sort(key=_sort_key)

        for r, snapshot in zip(sortable_positions, snapshots):
            _restore_row(ws, r, snapshot)


def update_excel(input_path, issues):
    if not os.path.exists(input_path):
        raise FileNotFoundError(
            f"Required file not found: '{input_path}'. "
            "Ensure 'Leahi Sprint Management.xlsx' exists in the working directory."
        )

    _, orig_table_xml = _capture_table_xml(input_path)

    wb = load_workbook(input_path)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Required sheet '{SHEET_NAME}' not found in '{input_path}'. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[SHEET_NAME]

    _ensure_sheet_setup(ws)
    _ensure_selection_sprints(wb)
    _ensure_data_validations(ws)

    current_sprint_num = _determine_current_sprint_num(issues, ws)

    # Build sprint_map: sprint_num → sprint_obj (for startDate/state checks in REMOVE).
    sprint_map: dict = {}
    for issue in issues:
        for s in (issue["fields"].get("customfield_10020") or []):
            num = _sprint_num(s.get("name", ""))
            if num >= MIN_SPRINT and num not in sprint_map:
                sprint_map[num] = s

    # Build api_map for the current sprint only.
    api_map = {}
    sys_excluded = set()
    for issue in issues:
        for sprint_num, scope in get_sprint_history(issue, current_sprint_num):
            if sprint_num != current_sprint_num or scope == "Move out of Sprint":
                continue
            composite = (issue["key"], sprint_num)
            sprint_obj = next(
                (s for s in (issue["fields"].get("customfield_10020") or [])
                 if _sprint_num(s.get("name", "")) == sprint_num),
                None
            )
            original_assignee = _original_assignee_for_sprint(issue, sprint_obj)
            team_name, team_idx = _team_for_nickname(original_assignee)
            if team_idx == _SYS_TEAM_IDX:
                sys_excluded.add(composite)
                continue
            api_map[composite] = {
                "issue":             issue,
                "sprint_num":        sprint_num,
                "scope":             scope,
                "jira_note":         issue.get("_jira_note"),
                "original_assignee": original_assignee,
                "team":              team_name,
            }

    existing = {}  
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        label      = ws.cell(row=row_idx, column=8).value
        sprint_val = ws.cell(row=row_idx, column=1).value
        key        = _extract_key(label)
        try:
            sprint_num = int(sprint_val) if sprint_val is not None else None
        except (ValueError, TypeError):
            sprint_num = None
        if key and sprint_num is not None and sprint_num == current_sprint_num:
            existing.setdefault((key, sprint_num), []).append(row_idx)

    updated = set()
    for composite, row_indices in existing.items():
        if composite in api_map:
            row_idx = row_indices[0]
            if ws.cell(row=row_idx, column=6).value == "Move out of Sprint":
                updated.add(composite)
                continue
            _write_issue_row(ws, row_idx, api_map[composite])
            updated.add(composite)

    rows_to_delete_unstarted = []
    rows_to_delete_sys = []
    removed_count = 0
    for composite, row_indices in existing.items():
        if composite in api_map:
            continue
        row_idx = row_indices[0]
        if composite in sys_excluded:
            if ws.cell(row=row_idx, column=6).value != "Move out of Sprint":
                rows_to_delete_sys.append(row_idx)
            continue
        sprint_num = composite[1]
        if not _sprint_started(sprint_num, sprint_map):
            rows_to_delete_unstarted.append(row_idx)
        elif ws.cell(row=row_idx, column=6).value != "Move out of Sprint":
            ws.cell(row=row_idx, column=6).value = "Move out of Sprint"
            removed_count += 1

    to_delete = list(rows_to_delete_unstarted) + list(rows_to_delete_sys)
    for row_indices in existing.values():
        to_delete.extend(row_indices[1:])
    to_delete.sort(reverse=True)
    for row_idx in to_delete:
        ws.delete_rows(row_idx)

    new_composites = [c for c in api_map if c not in existing]
    if new_composites:
        insert_after = _insert_point_for_sprint(ws, current_sprint_num)
        ws.insert_rows(insert_after + 1, len(new_composites))
        for j, composite in enumerate(new_composites):
            _write_issue_row(ws, insert_after + 1 + j, api_map[composite])

    for row_idx in range(ws.max_row, DATA_START_ROW - 1, -1):
        if any(ws.cell(row=row_idx, column=c).value is not None for c in range(1, 10)):
            break
        ws.delete_rows(row_idx)

    _sort_sprint_blocks(ws, current_sprint_num)

    final_row = ws.max_row

    if final_row >= DATA_START_ROW:
        _ensure_table_in_model(ws, f"A3:L{final_row}")

    col_names = [
        str(v) if (v := ws.cell(row=3, column=c).value) and str(v).strip() else f"Column{c}"
        for c in range(1, 13)
    ]

    _ensure_status_validation(ws)
    wb.save(input_path)

    _patch_xlsx_table(input_path, final_row, orig_table_xml, col_names)

    print(
        f"Saved: {input_path} "
        f"(current sprint: {current_sprint_num} | "
        f"{len(api_map)} active sprint-rows across {len(issues)} tickets | "
        f"{len(updated)} updated, {len(new_composites)} added, "
        f"{removed_count} marked removed, {len(rows_to_delete_unstarted)} deleted (unstarted sprint), "
        f"{len(rows_to_delete_sys)} deleted (SYS-team original assignee))"
    )
    return current_sprint_num


def _is_file_locked(path: str) -> bool:
    try:
        with open(path, "r+b"):
            return False
    except (PermissionError, OSError):
        return True


def _sync_hplus(source_path: str, dest_path: str) -> None:
    wb_src = load_workbook(source_path, data_only=True)
    ws_src = wb_src[SHEET_NAME]
    hplus_map = {}
    for row_idx in range(DATA_START_ROW, ws_src.max_row + 1):
        key = _extract_key(ws_src.cell(row=row_idx, column=8).value)
        try:
            sprint_num = int(ws_src.cell(row=row_idx, column=1).value)
        except (TypeError, ValueError):
            sprint_num = None
        if key and sprint_num is not None:
            hplus_map[(key, sprint_num)] = tuple(
                ws_src.cell(row=row_idx, column=c).value for c in SP_SYNC_COLS
            )

    _, dest_table_xml = _capture_table_xml(dest_path)

    wb_dst = load_workbook(dest_path)
    ws_dst = wb_dst[SHEET_NAME]
    for row_idx in range(DATA_START_ROW, ws_dst.max_row + 1):
        key = _extract_key(ws_dst.cell(row=row_idx, column=8).value)
        try:
            sprint_num = int(ws_dst.cell(row=row_idx, column=1).value)
        except (TypeError, ValueError):
            sprint_num = None
        if key and sprint_num is not None:
            vals = hplus_map.get((key, sprint_num))
            if vals is not None:
                for col_idx, val in zip(SP_SYNC_COLS, vals):
                    c            = ws_dst.cell(row=row_idx, column=col_idx, value=val)
                    c.protection = Protection(locked=False)

    final_row = ws_dst.max_row
    col_names = [
        str(v) if (v := ws_dst.cell(row=3, column=c).value) and str(v).strip() else f"Column{c}"
        for c in range(1, 13)
    ]
    wb_dst.save(dest_path)
    _patch_xlsx_table(dest_path, final_row, dest_table_xml, col_names)
    print(f"J/L notes synced from SharePoint -> X: drive ({len(hplus_map)} rows read)")


def _archive_excel(source_path: str, archive_dir: str) -> None:
    if not os.path.isdir(archive_dir):
        raise RuntimeError(
            f"ARCHIVE_PATH does not exist or is not a directory: '{archive_dir}'. "
            "Create the directory or update ARCHIVE_PATH in .env."
        )
    stem, ext = os.path.splitext(os.path.basename(source_path))
    date_suffix = datetime.now().strftime("%m%d%Y")
    archive_dest = os.path.join(archive_dir, f"{stem}_{date_suffix}{ext}")
    shutil.copy2(source_path, archive_dest)
    print(f"Archived: {archive_dest}")


def main():
    parser = argparse.ArgumentParser(
        description="Update Leahi Sprint Management Excel from Jira."
    )
    parser.add_argument(
        "input",
        nargs="?",
        default=os.getenv("EXCEL_PATH"),
        help="Path to the Excel file",
    )
    args = parser.parse_args()

    if not SHAREPOINT_SYNC_PATH:
        raise RuntimeError(
            "SHAREPOINT_SYNC_PATH is not set in .env. "
            "Set it to the Linux path of your OneDrive-synced SharePoint file."
        )

    if _is_file_locked(SHAREPOINT_SYNC_PATH):
        raise RuntimeError(
            "SharePoint file is currently open or locked — close it in Excel and re-run."
        )

    if not ARCHIVE_PATH:
        raise RuntimeError(
            "ARCHIVE_PATH is not set in .env. "
            "Set it to the directory where archive copies should be saved."
        )

    print("Fetching issues from Jira...")
    issues = get_jira_issues()
    print(f"Found {len(issues)} issues")
    print("Archiving Excel file...")
    _archive_excel(args.input, ARCHIVE_PATH)
    print("Updating Sprint Template on X: drive...")
    update_excel(args.input, issues)
    print("Syncing H-J notes from SharePoint to X: drive...")
    _sync_hplus(SHAREPOINT_SYNC_PATH, args.input)


if __name__ == "__main__":
    main()
