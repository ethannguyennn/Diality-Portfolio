"""Update Leahi Sprint Management on SharePoint from Jira API."""

import copy
import os
import re
import uuid
import zipfile
import requests
from requests.auth import HTTPBasicAuth
from dotenv import load_dotenv
from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.styles import Font, Alignment, Protection
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList
from datetime import datetime
from io import BytesIO

from sharepoint_helper import SharePointHelper
from config import SHAREPOINT_SYNC_PATH, ARCHIVE_PATH, SPRINT_TEMPLATE_FILE

load_dotenv()

# ── Jira ──
BASE_URL = "https://diality.atlassian.net"
PROJECT = "LDT"

# ── Sheet layout ──
SHEET_NAME = "Sprint Template"
DATA_START_ROW = 4
SHEET_HEADERS = [
    "Sprint", "Team", "Original Assignee", "Current Assignee",
    " Type", "Scope Changes", "Description", "Jira #", " Status",
]

# ── Teams ──
SWVV_TEAM = [
    "Raghu Kallala", "Thomas Lippold", "Tejaskumar Patel", "Tiffany Mejia",
    "Zoltan Miskolci", "Sarina Cheung", "Tisha Patel", "Ethan Nguyen",
]
FW_TEAM = [
    "Arpita Srivastava", "Jashwant Gantyada", "Michael Garthwaite", "Praneeth Bunne",
    "Sameer Poyil", "Varshini Nagabooshanam", "Vijay Pamula", "Suresh Dharnala",
    "Santhos Kumar Reddy", "Vinayakam Mani", "Sean Nash",
]
SW_TEAM = ["Nicholas Ramirez", "Stephen Quong", "Dara Navaei", "Behrouz NematiPour"]
SYS_TEAM = [
    "Eliza Petersen", "Caitlynn Chang", "Christina Heine", "Abhijit Barman",
    "Chris Yu", "Vitas Buenaventura", "Emiline Hernandez",
]
TEAMS = [FW_TEAM, SWVV_TEAM, SW_TEAM, SYS_TEAM]
TEAM_NAMES = ["FW", "SWVV", "SW", "SYS"]
_SYS_TEAM_IDX = len(TEAMS) - 1
_UNMATCHED_TEAM_IDX = len(TEAMS)

NICKNAME_MAP = {
    "Tejaskumar Patel": "Tejas", "Nicholas Ramirez": "Nico",
    "Thomas Lippold": "Tom", "Jashwant Gantyada": "Jashwant",
    "Behrouz NematiPour": "Behrouz", "Sean Nash": "Sean",
    "Vinayakam Mani": "Vinay", "Raghu Kallala": "Raghu",
    "Tiffany Mejia": "Tiffany", "Zoltan Miskolci": "Zoltan",
    "Sarina Cheung": "Sarina", "Tisha Patel": "Tisha",
    "Ethan Nguyen": "Ethan", "Arpita Srivastava": "Arpita",
    "Michael Garthwaite": "Michael", "Praneeth Bunne": "Praneeth",
    "Sameer Poyil": "Sameer", "Varshini Nagabooshanam": "Varshini",
    "Vijay Pamula": "Vijay", "Suresh Dharnala": "Suresh",
    "Dara Navaei": "Dara", "Eliza Petersen": "Eliza",
    "Stephen Quong": "Stephen", "Christina Heine": "Christina",
    "Caitlynn Chang": "Caitlynn", "Santhos Kumar Reddy": "Santhos",
    "Abhijit Barman": "Abhijit", "Chris Yu": "Chris",
    "Vitas Buenaventura": "Vitas", "Emiline Hernandez": "Emiline",
}

# Lookup: nickname or full name → team index
_ASSIGNEE_TO_TEAM_IDX: dict = {}
for _ti, _team in enumerate(TEAMS):
    for _full in _team:
        _ASSIGNEE_TO_TEAM_IDX[_full] = _ti
        _nick = NICKNAME_MAP.get(_full)
        if _nick:
            _ASSIGNEE_TO_TEAM_IDX[_nick] = _ti

# ── Excel table / validation constants ──
TABLE_DISPLAY_NAME = "Table1"
TABLE_STYLE = "TableStyleMedium7"
TYPE_FONT_COLORS = {
    "Bug": "FF0000", "Dialin ticket": "4472C4",
    "Support": "00B050", "Little-V": "7030A0", "Big-V": "000000",
}
BUG_DIALIN_TYPES = {"Bug", "Dialin ticket"}
BUG_STATUS_MAP = {
    "More Info Needed": "More Info Needed",
    "On Hold": "On Hold",
    "Blocked": "Blocked",
    "Ready for SW/FW Implementation": "Ready for SW/FW Implementation",
    "Investigation In Progress (SW/FW)": "Investigation in Progress",
    "Fix In Progress": "Fix in Progress",
    "Ready for Fix to be Tested": "Ready for Fix to be Tested",
    "Testing Fix in Progress": "Testing Fix in Progress",
    "Test Passed - Ready for Updated Release": "Ready for Updated Release",
    "Test Failed - SW/FW Fix Needed": "SW/FW Fix Needed",
    "Test Passed - Needs Final Fix Details": "Needs Final Fix Details",
    "Done": "Done",
}
MIN_SPRINT = 28
SP_SYNC_COLS = (10, 12)
JIRA_NOTE_RE = re.compile(r'xlsx<([^>]*)>xlsx')
VALIDATION_ROW_END = 1000
_SELECTION_SPRINT_MAX = 100
_COL_VALIDATIONS = {
    "A": f"Selection!$F$2:$F${_SELECTION_SPRINT_MAX + 1}",
    "B": "Selection!$B$2:$B$5",
    "C": "Selection!$A$2:$A$27",
    "D": "Selection!$A$2:$A$27",
    "E": "Selection!$D$2:$D$6",
    "F": "Selection!$H$2:$H$4",
    "G": "Selection!$K$2:$K$3",
    "J": "Selection!$J$2:$J$5",
}
_SORT_NUM_COLS = 12


# ═══════════════════════════════════════════════════════════════════
#  Jira helpers
# ═══════════════════════════════════════════════════════════════════

def _fetch_paginated(url, jql, auth):
    all_issues, start_at = [], 0
    while True:
        params = {
            "fields": "summary,status,assignee,issuetype,customfield_10020,description,comment",
            "expand": "changelog", "jql": jql, "maxResults": 100, "startAt": start_at,
        }
        resp = requests.get(url, params=params, auth=auth)
        resp.raise_for_status()
        data = resp.json()
        all_issues.extend(data["issues"])
        start_at += len(data["issues"])
        if start_at >= data.get("total", 0) or not data["issues"]:
            break
    return all_issues


def _sprint_names(s):
    return {n.strip() for n in s.split(",") if n.strip()} if s else set()


def _sprint_num(name):
    m = re.search(r'(\d+)', name)
    return int(m.group(1)) if m else -1


def _has_sprint_history(issue):
    for history in issue["changelog"]["histories"]:
        for item in history["items"]:
            if item["field"] != "Sprint":
                continue
            for name in _sprint_names(item.get("toString") or ""):
                if _sprint_num(name) >= MIN_SPRINT:
                    return True
    return False


def _assignee_nickname(display_name: str) -> str:
    if not display_name:
        return "Unassigned"
    return NICKNAME_MAP.get(display_name, display_name.split()[0])


def _team_for_nickname(nickname):
    idx = _ASSIGNEE_TO_TEAM_IDX.get(nickname)
    if idx is None:
        return None, _UNMATCHED_TEAM_IDX
    return TEAM_NAMES[idx], idx


def _adf_to_text(node) -> str:
    if isinstance(node, str):
        return node
    if not isinstance(node, dict):
        return ""
    if node.get("type") == "text":
        return node.get("text", "")
    return "".join(_adf_to_text(c) for c in node.get("content", []))


def _latest_jira_note(issue, auth) -> str:
    comment_field = issue["fields"].get("comment") or {}
    comments = list(comment_field.get("comments", []))
    if comment_field.get("total", len(comments)) > len(comments):
        url = f"{BASE_URL}/rest/api/2/issue/{issue['key']}/comment"
        comments = requests.get(url, params={"maxResults": 1000}, auth=auth).json().get("comments", [])

    best_text, best_created = None, ""
    for comment in comments:
        body_text = _adf_to_text(comment.get("body") or "")
        m = JIRA_NOTE_RE.search(body_text)
        if m and comment.get("created", "") >= best_created:
            best_created = comment["created"]
            best_text = m.group(1).strip()
    return best_text


def get_jira_issues():
    auth = HTTPBasicAuth(os.getenv("JIRA_EMAIL"), os.getenv("JIRA_API_TOKEN"))

    board_issues = _fetch_paginated(
        f"{BASE_URL}/rest/agile/1.0/board/84/issue",
        "sprint in openSprints() OR sprint in futureSprints()", auth,
    )
    seen = {i["key"] for i in board_issues}

    all_backlog = _fetch_paginated(
        f"{BASE_URL}/rest/agile/1.0/board/84/backlog",
        f"project = {PROJECT}", auth,
    )
    backlog_with_history = [
        i for i in all_backlog if i["key"] not in seen and _has_sprint_history(i)
    ]

    all_issues = board_issues + backlog_with_history
    for issue in all_issues:
        issue["_jira_note"] = _latest_jira_note(issue, auth)
    return all_issues


def parse_date(date_str):
    if not date_str:
        return None
    date_str = date_str.replace("Z", "+00:00")
    date_str = re.sub(r'([+-])(\d{2})(\d{2})$', r'\1\2:\3', date_str)
    return datetime.fromisoformat(date_str)


def get_type(issue):
    issuetype = issue["fields"]["issuetype"]["name"].lower()
    summary = issue["fields"].get("summary", "").lower()
    if "dialin" in issuetype or "dialin" in summary:
        return "Dialin ticket"
    if issuetype == "bug":
        return "Bug"
    if "little" in issuetype:
        return "Little-V"
    if "support" in summary:
        return "Support"
    return "Big-V"


def get_status(issue, issue_type) -> str:
    status = issue["fields"]["status"]
    status_name = status["name"]
    if issue_type in BUG_DIALIN_TYPES:
        return BUG_STATUS_MAP.get(status_name, status_name)

    status_name = status_name.lower()
    category_key = status["statusCategory"]["key"]
    if "blocked" in status_name:
        return "Blocked"
    if "on hold" in status_name or "on-hold" in status_name:
        return "On Hold"
    if status_name == "n/a":
        return "N/A"
    if category_key == "done" or "done" in status_name or "complete" in status_name:
        return "Done"
    if category_key == "new":
        return "Not started"
    return "In-progress"


# ═══════════════════════════════════════════════════════════════════
#  Sprint history / timing
# ═══════════════════════════════════════════════════════════════════

def _classify_sprint_timing(sprint, histories):
    sprint_name = sprint["name"]
    sprint_start_dt = parse_date(sprint.get("startDate"))
    if not sprint_start_dt:
        return "Scoped"

    most_recent_add = None
    for history in histories:
        for item in history["items"]:
            if item["field"] != "Sprint":
                continue
            to_sprints = _sprint_names(item.get("toString") or "")
            from_sprints = _sprint_names(item.get("fromString") or "")
            if sprint_name in to_sprints and sprint_name not in from_sprints:
                added_dt = parse_date(history["created"])
                if added_dt and (most_recent_add is None or added_dt > most_recent_add):
                    most_recent_add = added_dt

    if most_recent_add is None:
        return "Scoped"
    sprint_start_eod = sprint_start_dt.replace(hour=23, minute=59, second=59, microsecond=0)
    return "Added" if most_recent_add > sprint_start_eod else "Scoped"


def _determine_current_sprint_num(issues, ws) -> int:
    active = []
    for issue in issues:
        for s in (issue["fields"].get("customfield_10020") or []):
            num = _sprint_num(s.get("name", ""))
            if num >= MIN_SPRINT and s.get("state") == "active":
                active.append(num)
    if active:
        return max(active)
    try:
        b2 = int(ws.cell(row=2, column=2).value)
        if b2 >= MIN_SPRINT:
            return b2
    except (TypeError, ValueError):
        pass
    return MIN_SPRINT


def get_sprint_history(issue, max_sprint_num):
    histories = issue["changelog"]["histories"]
    sprint_info = issue["fields"].get("customfield_10020") or []
    current_sprint_names = {s["name"] for s in sprint_info}
    current_sprint_objs = {s["name"]: s for s in sprint_info}

    all_sprint_names = set(current_sprint_names)
    for history in histories:
        for item in history["items"]:
            if item["field"] != "Sprint":
                continue
            for name in _sprint_names(item.get("fromString") or ""):
                all_sprint_names.add(name)
            for name in _sprint_names(item.get("toString") or ""):
                all_sprint_names.add(name)

    sprint_entries = {}
    for name in all_sprint_names:
        if name not in current_sprint_names:
            num = _sprint_num(name)
            if num >= MIN_SPRINT and num <= max_sprint_num:
                sprint_entries[num] = "Move out of Sprint"

    for name, sprint_obj in current_sprint_objs.items():
        num = _sprint_num(name)
        if num >= MIN_SPRINT:
            sprint_entries[num] = _classify_sprint_timing(sprint_obj, histories)

    return sorted(sprint_entries.items(), key=lambda x: x[0])


def _original_assignee_for_sprint(issue, sprint_obj) -> str:
    # Assignee at sprint entry (added) or EOD of sprint start (scoped).
    fields = issue["fields"]
    histories = issue["changelog"]["histories"]
    current_full = fields["assignee"]["displayName"] if fields.get("assignee") else ""

    if not sprint_obj:
        return _assignee_nickname(current_full)

    sprint_name = sprint_obj.get("name", "")
    sprint_added_dt = None
    for history in histories:
        for item in history["items"]:
            if item["field"] != "Sprint":
                continue
            to_sprints = _sprint_names(item.get("toString") or "")
            from_sprints = _sprint_names(item.get("fromString") or "")
            if sprint_name in to_sprints and sprint_name not in from_sprints:
                dt = parse_date(history["created"])
                if sprint_added_dt is None or dt < sprint_added_dt:
                    sprint_added_dt = dt

    if sprint_added_dt is None:
        return _assignee_nickname(current_full)

    sprint_start_dt = parse_date(sprint_obj.get("startDate"))
    if sprint_start_dt:
        sprint_start_eod = sprint_start_dt.replace(hour=23, minute=59, second=59, microsecond=0)
        cutoff_dt = max(sprint_added_dt, sprint_start_eod)
    else:
        cutoff_dt = sprint_added_dt

    assignee_changes = []
    for history in histories:
        for item in history["items"]:
            if item["field"] == "assignee":
                assignee_changes.append((parse_date(history["created"]), item))
    assignee_changes.sort(key=lambda x: x[0])

    if not assignee_changes:
        return _assignee_nickname(current_full)

    current_state = assignee_changes[0][1].get("fromString") or ""
    for change_dt, item in assignee_changes:
        if change_dt <= cutoff_dt:
            current_state = item.get("toString") or ""
        else:
            break
    return _assignee_nickname(current_state) if current_state else "Unassigned"


# ═══════════════════════════════════════════════════════════════════
#  Excel sheet setup / validation
# ═══════════════════════════════════════════════════════════════════

def _extract_key(label):
    m = re.match(r'\[([A-Z]+-\d+)\]', str(label) if label is not None else '')
    return m.group(1) if m else None


def _ensure_sheet_setup(ws):
    for col, header in enumerate(SHEET_HEADERS, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 18.75
    ws.freeze_panes = "A4"
    for letter, width in {"A": 9.29, "B": 9, "C": 10.57, "D": 10.57, "E": 11.57, "F": 18.43, "H": 103.14, "I": 16.29}.items():
        ws.column_dimensions[letter].width = width


def _ensure_selection_sprints(wb):
    if "Selection" not in wb.sheetnames:
        return
    ws_sel = wb["Selection"]
    for sprint_num in range(1, _SELECTION_SPRINT_MAX + 1):
        row = sprint_num + 1
        if ws_sel.cell(row=row, column=6).value is None:
            ws_sel.cell(row=row, column=6).value = sprint_num


def _ensure_data_validations(ws):
    ws.data_validations = DataValidationList()
    for col_letter, formula in _COL_VALIDATIONS.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                            showErrorMessage=True, errorTitle="Invalid entry",
                            error="Please select from the dropdown list.", errorStyle="warning")
        dv.add(f"{col_letter}{DATA_START_ROW}:{col_letter}{VALIDATION_ROW_END}")
        ws.add_data_validation(dv)


def _ensure_status_validation(ws):
    _L_RANGE = "Selection!$L$2:$L$13"
    _E_RANGE = "Selection!$E$2:$E$7"

    row_types = [
        (r, ws.cell(row=r, column=5).value in BUG_DIALIN_TYPES)
        for r in range(DATA_START_ROW, ws.max_row + 1)
    ]

    groups = []
    if row_types:
        cur_start, cur_type = row_types[0]
        for row_idx, is_bug in row_types[1:]:
            if is_bug != cur_type:
                groups.append((cur_start, row_idx - 1, cur_type))
                cur_start, cur_type = row_idx, is_bug
        groups.append((cur_start, row_types[-1][0], cur_type))
        if row_types[-1][0] < VALIDATION_ROW_END:
            groups.append((row_types[-1][0] + 1, VALIDATION_ROW_END, False))
    else:
        groups.append((DATA_START_ROW, VALIDATION_ROW_END, False))

    for start_row, end_row, is_bug_dialin in groups:
        dv = DataValidation(type="list", formula1=_L_RANGE if is_bug_dialin else _E_RANGE,
                            allow_blank=True, showErrorMessage=True, errorTitle="Invalid entry",
                            error="Please select from the dropdown list.", errorStyle="warning")
        dv.add(f"I{start_row}:I{end_row}")
        ws.add_data_validation(dv)


# ═══════════════════════════════════════════════════════════════════
#  Row writing / table management
# ═══════════════════════════════════════════════════════════════════

def _write_issue_row(ws, row_idx, row_record):
    issue = row_record["issue"]
    fields = issue["fields"]
    key = issue["key"]

    full_name = fields["assignee"]["displayName"] if fields.get("assignee") else ""
    current_assignee = _assignee_nickname(full_name)
    original_assignee = row_record.get("original_assignee") or current_assignee
    issue_type = get_type(issue)
    ticket_url = f"{BASE_URL}/browse/{key}"
    ticket_label = f"[{key}] {fields.get('summary', '')}"
    has_description = "Y" if len(_adf_to_text(fields.get("description") or "")) > 20 else "N"
    status = get_status(issue, issue_type)

    values = [
        row_record["sprint_num"], row_record.get("team"), original_assignee,
        current_assignee, issue_type, row_record["scope"], has_description,
        ticket_label, status,
    ]
    align = Alignment(horizontal="left", vertical="center")

    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.alignment = align
        cell.protection = Protection(locked=False)
        if col == 8:
            cell.font = Font(name="Calibri", size=11, color="1155CC", underline="single")
            cell.hyperlink = ticket_url
        elif col == 5:
            cell.font = Font(name="Calibri", size=11, color=TYPE_FONT_COLORS.get(issue_type, "000000"))
            cell.hyperlink = None
        else:
            cell.hyperlink = None

    jira_note = row_record.get("jira_note")
    if jira_note is not None:
        c = ws.cell(row=row_idx, column=11, value=jira_note)
        c.alignment = align
        c.protection = Protection(locked=False)


def _ensure_table_in_model(ws, new_ref: str):
    if ws.tables:
        for tbl in ws.tables.values():
            tbl.ref = new_ref
    else:
        tbl = Table(displayName=TABLE_DISPLAY_NAME, ref=new_ref)
        tbl.tableStyleInfo = TableStyleInfo(name=TABLE_STYLE, showRowStripes=True)
        min_col, _, max_col, _ = range_boundaries(new_ref)
        tbl.tableColumns = [
            TableColumn(
                id=i,
                name=str(v) if (v := ws.cell(row=3, column=col).value) and str(v).strip() else f"Column{col}"
            )
            for i, col in enumerate(range(min_col, max_col + 1), start=1)
        ]
        ws.add_table(tbl)


# ═══════════════════════════════════════════════════════════════════
#  Table XML patching (in-memory, preserves VBA entries)
# ═══════════════════════════════════════════════════════════════════

def _xml_attr(value: str) -> str:
    return value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _generate_table_xml(ref, col_names, uid=None):
    xr_uid = uid or ("{" + str(uuid.uuid4()).upper() + "}")
    cols_xml = "".join(f'<tableColumn id="{i+1}" name="{_xml_attr(n)}"/>' for i, n in enumerate(col_names))
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<table xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        ' xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"'
        ' xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision"'
        ' xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"'
        ' mc:Ignorable="xr xr3"'
        f' id="1" name="{TABLE_DISPLAY_NAME}" displayName="{TABLE_DISPLAY_NAME}"'
        f' ref="{ref}" totalsRowShown="0" xr:uid="{xr_uid}">'
        f'<autoFilter ref="{ref}"/>'
        f'<tableColumns count="{len(col_names)}">{cols_xml}</tableColumns>'
        f'<tableStyleInfo name="{TABLE_STYLE}" showFirstColumn="0" showLastColumn="0"'
        ' showRowStripes="1" showColumnStripes="0"/>'
        '</table>'
    ).encode("utf-8")


def _capture_table_xml_from_bytes(wb_bytes):
    with zipfile.ZipFile(BytesIO(wb_bytes), "r") as z:
        table_paths = [n for n in z.namelist() if n.startswith("xl/tables/") and n.endswith(".xml")]
        if not table_paths:
            return None, None
        path = table_paths[0]
        return path, z.read(path).decode("utf-8", errors="replace")


def _patch_table_xml_in_bytes(wb_bytes, final_row, orig_table_xml=None, col_names=None):
    ref = f"A3:L{final_row}"
    with zipfile.ZipFile(BytesIO(wb_bytes), "r") as zin:
        table_paths = [n for n in zin.namelist() if n.startswith("xl/tables/") and n.endswith(".xml")]
        if not table_paths:
            return wb_bytes
        table_path = table_paths[0]
        raw_xml = zin.read(table_path).decode("utf-8", errors="replace")
        entries = [(info, zin.read(info.filename)) for info in zin.infolist()]

    source = orig_table_xml or raw_xml
    uid_m = re.search(r'xr:uid="([^"]*)"', source)
    clean_xml = _generate_table_xml(ref, col_names or [], uid=uid_m.group(1) if uid_m else None)

    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for info, data in entries:
            if info.filename == table_path:
                zout.writestr(info.filename, clean_xml)
            else:
                zout.writestr(info, data)  # preserves ZipInfo metadata (VBA, etc.)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════
#  Row sorting / sprint positioning
# ═══════════════════════════════════════════════════════════════════

def _insert_point_for_sprint(ws, sprint_num):
    result = DATA_START_ROW - 1
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        try:
            n = int(ws.cell(row=row_idx, column=1).value)
        except (TypeError, ValueError):
            continue
        if n <= sprint_num:
            result = row_idx
    return result


def _sprint_started(sprint_num, sprint_map):
    s = sprint_map.get(sprint_num)
    if not s:
        return True
    state = s.get("state", "")
    if state in ("active", "closed"):
        return True
    if state == "future":
        return False
    start_dt = parse_date(s.get("startDate"))
    return datetime.now(start_dt.tzinfo) >= start_dt if start_dt else False


def _snapshot_row(ws, row_idx, num_cols=_SORT_NUM_COLS):
    cells = []
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cells.append({
            "value": cell.value, "font": copy.copy(cell.font),
            "fill": copy.copy(cell.fill), "border": copy.copy(cell.border),
            "alignment": copy.copy(cell.alignment), "number_format": cell.number_format,
            "protection": copy.copy(cell.protection), "hyperlink": copy.copy(cell.hyperlink),
            "comment": copy.copy(cell.comment),
        })
    return cells


def _restore_row(ws, row_idx, snapshot):
    for col, state in enumerate(snapshot, start=1):
        cell = ws.cell(row=row_idx, column=col)
        for attr in ("value", "font", "fill", "border", "alignment", "number_format", "protection", "hyperlink", "comment"):
            setattr(cell, attr, state[attr])


def _sort_sprint_blocks(ws, target_sprints):
    blocks = []
    row_idx = DATA_START_ROW
    while row_idx <= ws.max_row:
        try:
            sprint_num = int(ws.cell(row=row_idx, column=1).value)
        except (TypeError, ValueError):
            row_idx += 1
            continue
        end_row = row_idx
        while end_row + 1 <= ws.max_row:
            try:
                if int(ws.cell(row=end_row + 1, column=1).value) != sprint_num:
                    break
            except (TypeError, ValueError):
                break
            end_row += 1
        blocks.append((row_idx, end_row, sprint_num))
        row_idx = end_row + 1

    def _sort_key(snap):
        _, team_idx = _team_for_nickname(snap[2]["value"])
        return (team_idx, snap[2]["value"] or "")

    for start_row, end_row, sprint_num in blocks:
        if sprint_num not in target_sprints:
            continue
        positions, snaps = [], []
        for r in range(start_row, end_row + 1):
            if ws.cell(row=r, column=6).value == "Move out of Sprint":
                continue
            positions.append(r)
            snaps.append(_snapshot_row(ws, r))
        if len(positions) <= 1:
            continue
        snaps.sort(key=_sort_key)
        for r, snap in zip(positions, snaps):
            _restore_row(ws, r, snap)


# ═══════════════════════════════════════════════════════════════════
#  Manual notes (J/L columns) capture / restore
# ═══════════════════════════════════════════════════════════════════

def _capture_manual_notes(ws):
    notes = {}
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        key = _extract_key(ws.cell(row=row_idx, column=8).value)
        try:
            sprint_num = int(ws.cell(row=row_idx, column=1).value)
        except (TypeError, ValueError):
            continue
        if key:
            notes[(key, sprint_num)] = tuple(ws.cell(row=row_idx, column=c).value for c in SP_SYNC_COLS)
    return notes


def _restore_manual_notes(ws, notes):
    restored = 0
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        key = _extract_key(ws.cell(row=row_idx, column=8).value)
        try:
            sprint_num = int(ws.cell(row=row_idx, column=1).value)
        except (TypeError, ValueError):
            continue
        if key:
            vals = notes.get((key, sprint_num))
            if vals is not None:
                for col_idx, val in zip(SP_SYNC_COLS, vals):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.protection = Protection(locked=False)
                restored += 1
    return restored


# ═══════════════════════════════════════════════════════════════════
#  Core update logic
# ═══════════════════════════════════════════════════════════════════

def update_sharepoint(helper, template_path, sync_path, issues):
    # Read manual notes from sync file (read-only)
    print(f"Reading manual notes from: {sync_path}")
    sync_bytes = helper.read_excel(sync_path)
    wb_sync = openpyxl_load_workbook(BytesIO(sync_bytes), data_only=True)
    manual_notes = _capture_manual_notes(wb_sync[SHEET_NAME])
    print(f"Captured {len(manual_notes)} rows of manual notes (columns J/L)")

    # Read template file
    print(f"Reading template: {template_path}")
    template_bytes = helper.read_excel(template_path)
    _, orig_table_xml = _capture_table_xml_from_bytes(template_bytes)

    is_xlsm = template_path.lower().endswith('.xlsm')
    wb = openpyxl_load_workbook(BytesIO(template_bytes), keep_vba=is_xlsm)
    ws = wb[SHEET_NAME]

    _ensure_sheet_setup(ws)
    _ensure_selection_sprints(wb)
    _ensure_data_validations(ws)

    current_sprint_num = _determine_current_sprint_num(issues, ws)
    next_sprint_num = current_sprint_num + 1
    target_sprints = (current_sprint_num, next_sprint_num)
    target_sprint_set = set(target_sprints)

    # Sprint object lookup
    sprint_map = {}
    for issue in issues:
        for s in (issue["fields"].get("customfield_10020") or []):
            num = _sprint_num(s.get("name", ""))
            if num >= MIN_SPRINT and num not in sprint_map:
                sprint_map[num] = s

    # Build api_map (active sprint rows)
    api_map, sys_excluded = {}, set()
    for issue in issues:
        for sprint_num, scope in get_sprint_history(issue, next_sprint_num):
            if sprint_num not in target_sprint_set or scope == "Move out of Sprint":
                continue
            composite = (issue["key"], sprint_num)
            sprint_obj = next(
                (s for s in (issue["fields"].get("customfield_10020") or [])
                 if _sprint_num(s.get("name", "")) == sprint_num), None
            )
            original_assignee = _original_assignee_for_sprint(issue, sprint_obj)
            current_full = issue["fields"]["assignee"]["displayName"] if issue["fields"].get("assignee") else ""
            current_assignee = _assignee_nickname(current_full)
            team_name, team_idx = _team_for_nickname(current_assignee)
            if team_idx == _SYS_TEAM_IDX:
                sys_excluded.add(composite)
                continue
            api_map[composite] = {
                "issue": issue, "sprint_num": sprint_num, "scope": scope,
                "jira_note": issue.get("_jira_note"),
                "original_assignee": original_assignee, "team": team_name,
            }

    # Index existing rows
    existing = {}
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        key = _extract_key(ws.cell(row=row_idx, column=8).value)
        try:
            sn = int(ws.cell(row=row_idx, column=1).value) if ws.cell(row=row_idx, column=1).value is not None else None
        except (ValueError, TypeError):
            sn = None
        if key and sn is not None and sn in target_sprint_set:
            existing.setdefault((key, sn), []).append(row_idx)

    # Update existing rows
    updated = set()
    for composite, rows in existing.items():
        if composite in api_map:
            if ws.cell(row=rows[0], column=6).value == "Move out of Sprint":
                updated.add(composite)
                continue
            _write_issue_row(ws, rows[0], api_map[composite])
            updated.add(composite)

    # Mark removed / delete
    rows_to_delete_unstarted, rows_to_delete_sys = [], []
    removed_count = 0
    for composite, rows in existing.items():
        if composite in api_map:
            continue
        row_idx = rows[0]
        if composite in sys_excluded:
            if ws.cell(row=row_idx, column=6).value != "Move out of Sprint":
                rows_to_delete_sys.append(row_idx)
            continue
        if not _sprint_started(composite[1], sprint_map):
            rows_to_delete_unstarted.append(row_idx)
        elif ws.cell(row=row_idx, column=6).value != "Move out of Sprint":
            ws.cell(row=row_idx, column=6).value = "Move out of Sprint"
            removed_count += 1

    to_delete = rows_to_delete_unstarted + rows_to_delete_sys
    for row_indices in existing.values():
        to_delete.extend(row_indices[1:])
    for row_idx in sorted(to_delete, reverse=True):
        ws.delete_rows(row_idx)

    # Insert new rows
    new_composites = [c for c in api_map if c not in existing]
    for sn in target_sprints:
        sprint_new = [c for c in new_composites if c[1] == sn]
        if not sprint_new:
            continue
        insert_after = _insert_point_for_sprint(ws, sn)
        ws.insert_rows(insert_after + 1, len(sprint_new))
        for j, composite in enumerate(sprint_new):
            _write_issue_row(ws, insert_after + 1 + j, api_map[composite])

    # Trim trailing empty rows
    for row_idx in range(ws.max_row, DATA_START_ROW - 1, -1):
        if any(ws.cell(row=row_idx, column=c).value is not None for c in range(1, 10)):
            break
        ws.delete_rows(row_idx)

    _sort_sprint_blocks(ws, target_sprint_set)

    # Restore J/L notes from sync file
    notes_restored = _restore_manual_notes(ws, manual_notes)
    print(f"Restored {notes_restored} rows of manual notes")

    final_row = ws.max_row
    if final_row >= DATA_START_ROW:
        _ensure_table_in_model(ws, f"A3:L{final_row}")

    col_names = [
        str(v) if (v := ws.cell(row=3, column=c).value) and str(v).strip() else f"Column{c}"
        for c in range(1, 13)
    ]
    _ensure_status_validation(ws)

    # Save to bytes → patch table XML → upload
    buf = BytesIO()
    wb.save(buf)
    wb_bytes = _patch_table_xml_in_bytes(buf.getvalue(), final_row, orig_table_xml, col_names)

    print(f"Uploading to SharePoint: {template_path}")
    helper.write_excel(template_path, wb_bytes)

    print(
        f"Done: {template_path} "
        f"(sprints {current_sprint_num}/{next_sprint_num} | "
        f"{len(api_map)} active rows, {len(issues)} tickets | "
        f"{len(updated)} updated, {len(new_composites)} added, "
        f"{removed_count} removed, {len(rows_to_delete_unstarted)} deleted-unstarted, "
        f"{len(rows_to_delete_sys)} deleted-SYS)"
    )
    return current_sprint_num


# ═══════════════════════════════════════════════════════════════════
#  Entry point
# ═══════════════════════════════════════════════════════════════════

def main():
    helper = SharePointHelper()

    print("Fetching issues from Jira...")
    issues = get_jira_issues()
    print(f"Found {len(issues)} issues")

    print(f"Archiving: {SPRINT_TEMPLATE_FILE}")
    archive_path = helper.archive_file(SPRINT_TEMPLATE_FILE, archive_folder=ARCHIVE_PATH)
    print(f"Archived: {archive_path}")

    print("Updating Sprint Template...")
    update_sharepoint(helper, SPRINT_TEMPLATE_FILE, SHAREPOINT_SYNC_PATH, issues)


if __name__ == "__main__":
    main()
