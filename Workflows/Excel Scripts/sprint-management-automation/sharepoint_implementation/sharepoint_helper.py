"""SharePoint Excel file operations via Microsoft Graph API."""

import logging
import os
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

from config import (
    AUTH_ENDPOINT,
    GRAPH_API_BASE,
    LOG_FORMAT,
    LOG_LEVEL,
    SHAREPOINT_DRIVE_ID,
    SHAREPOINT_SITE_ID,
)

load_dotenv()

logger = logging.getLogger(__name__)
logging.basicConfig(level=LOG_LEVEL, format=LOG_FORMAT)


class SharePointAuthError(Exception):
    pass


class SharePointOperationError(Exception):
    pass


def _short_id(value: str, length: int = 12) -> str:
    """Truncate a long opaque id (e.g. a WOPI session id) for log readability."""
    return value if len(value) <= length else f"{value[:length]}..."


class SharePointHelper:
    """Read/write Excel files on SharePoint via Microsoft Graph API."""

    def __init__(self, tenant_id=None, client_id=None, client_secret=None,
                 site_id=SHAREPOINT_SITE_ID, drive_id=SHAREPOINT_DRIVE_ID):
        self.tenant_id = tenant_id or os.getenv("TENANT_ID")
        self.client_id = client_id or os.getenv("CLIENT_ID")
        self.client_secret = client_secret or os.getenv("CLIENT_SECRET")
        self.site_id = site_id
        self.drive_id = drive_id

        if not all([self.tenant_id, self.client_id, self.client_secret]):
            logger.error("Missing credentials: TENANT_ID, CLIENT_ID, or CLIENT_SECRET not set in .env")
            raise SharePointAuthError(
                "Missing credentials. Set TENANT_ID, CLIENT_ID, CLIENT_SECRET in .env."
            )

        self.access_token = self._fetch_access_token()

    def _fetch_access_token(self) -> str:
        logger.info("Fetching new Graph API access token")
        url = AUTH_ENDPOINT.format(tenant_id=self.tenant_id)
        payload = {
            "grant_type": "client_credentials",
            "client_id": self.client_id,
            "client_secret": self.client_secret,
            "scope": "https://graph.microsoft.com/.default",
        }

        try:
            resp = requests.post(url, data=payload,
                                 headers={"Content-Type": "application/x-www-form-urlencoded"},
                                 timeout=10)
            resp.raise_for_status()
            data = resp.json()
            token = data.get("access_token")
            if not token:
                logger.error("Token endpoint returned no access_token")
                raise SharePointAuthError("No access token in response")
            logger.info("Access token acquired successfully")
            return token
        except requests.RequestException as e:
            logger.error(f"Authentication failed: {e}")
            raise SharePointAuthError(f"Authentication failed: {e}")

    def _get_headers(self) -> dict:
        return {"Authorization": f"Bearer {self.access_token}", "Content-Type": "application/json"}

    def _file_url(self, file_path: str) -> str:
        return f"{GRAPH_API_BASE}/sites/{self.site_id}/drives/{self.drive_id}/root:/{file_path}:/content"

    # ── Internal helpers ──

    def _drive_base(self) -> str:
        return f"{GRAPH_API_BASE}/sites/{self.site_id}/drives/{self.drive_id}"

    def get_item_id(self, file_path: str) -> str:
        """Resolve a file path to its driveItem id."""
        logger.debug(f"Resolving item id for {file_path}")
        url = f"{self._drive_base()}/root:/{file_path}"
        try:
            resp = requests.get(url, headers=self._get_headers(), timeout=15)
            resp.raise_for_status()
            item_id = resp.json()["id"]
            logger.debug(f"Resolved item id for {file_path}: {item_id}")
            return item_id
        except requests.RequestException as e:
            logger.error(f"Could not resolve item id for {file_path}: {e}")
            raise SharePointOperationError(f"Could not resolve item id for {file_path}: {e}")

    def get_item_id_if_exists(self, file_path: str) -> Optional[str]:
        """Resolve a path to its driveItem id, or None if it doesn't exist."""
        url = f"{self._drive_base()}/root:/{file_path}"
        try:
            resp = requests.get(url, headers=self._get_headers(), timeout=15)
            if resp.status_code == 404:
                return None
            resp.raise_for_status()
            return resp.json()["id"]
        except requests.RequestException as e:
            logger.error(f"Could not check existence of {file_path}: {e}")
            raise SharePointOperationError(f"Could not check existence of {file_path}: {e}")

    def list_children(self, folder_path: str) -> list:
        """List the immediate children (files and folders) of a SharePoint folder."""
        logger.debug(f"Listing children of {folder_path}")
        url = (f"{self._drive_base()}/root:/{folder_path}:/children"
               f"?$select=id,name,size,lastModifiedDateTime,folder,file&$top=999")
        items = []
        try:
            while url:
                resp = requests.get(url, headers=self._get_headers(), timeout=30)
                resp.raise_for_status()
                data = resp.json()
                items.extend(data.get("value", []))
                url = data.get("@odata.nextLink")
            logger.debug(f"Listed {len(items)} children of {folder_path}")
            return items
        except requests.RequestException as e:
            logger.error(f"Failed to list children of {folder_path}: {e}")
            raise SharePointOperationError(f"Failed to list children of {folder_path}: {e}")

    def find_file_path_in_tree(self, folder_path: str, filename: str) -> Optional[str]:
        """Recursively search folder_path and its subfolders for a file named
        filename. Returns the full path to the first match found, or None.
        """
        children = self.list_children(folder_path)
        subfolders = []
        for c in children:
            if "file" in c and c["name"] == filename:
                return f"{folder_path}/{filename}"
            if "folder" in c:
                subfolders.append(c["name"])
        for name in subfolders:
            found = self.find_file_path_in_tree(f"{folder_path}/{name}", filename)
            if found:
                return found
        return None

    def ensure_folder(self, folder_path: str) -> str:
        """Return the item id for folder_path, creating it if it doesn't exist.

        Only creates the leaf folder; the parent path must already exist.
        """
        existing = self.get_item_id_if_exists(folder_path)
        if existing:
            return existing

        parent_path = str(Path(folder_path).parent)
        folder_name = Path(folder_path).name
        url = (f"{self._drive_base()}/root/children" if parent_path == "."
               else f"{self._drive_base()}/root:/{parent_path}:/children")
        logger.info(f"Creating folder: {folder_path}")
        try:
            resp = requests.post(
                url,
                json={"name": folder_name, "folder": {}, "@microsoft.graph.conflictBehavior": "fail"},
                headers=self._get_headers(), timeout=30)
            resp.raise_for_status()
            item_id = resp.json()["id"]
            logger.info(f"Created folder: {folder_path}")
            return item_id
        except requests.RequestException as e:
            logger.error(f"Failed to create folder {folder_path}: {e}")
            raise SharePointOperationError(f"Failed to create folder {folder_path}: {e}")

    def move_item(self, item_id: str, dest_folder_id: str) -> None:
        """Move a file/folder to a new parent folder, keeping its name."""
        url = f"{self._drive_base()}/items/{item_id}"
        try:
            resp = requests.patch(
                url, json={"parentReference": {"id": dest_folder_id}},
                headers=self._get_headers(), timeout=30)
            resp.raise_for_status()
        except requests.RequestException as e:
            logger.error(f"Failed to move item {item_id} to folder {dest_folder_id}: {e}")
            raise SharePointOperationError(f"Failed to move item {item_id}: {e}")

    def delete_item(self, item_id: str) -> None:
        """Delete a file/folder (moves it to the SharePoint recycle bin)."""
        url = f"{self._drive_base()}/items/{item_id}"
        try:
            resp = requests.delete(url, headers=self._get_headers(), timeout=30)
            resp.raise_for_status()
        except requests.RequestException as e:
            logger.error(f"Failed to delete item {item_id}: {e}")
            raise SharePointOperationError(f"Failed to delete item {item_id}: {e}")

    # ── File I/O ──

    def read_excel(self, file_path: str) -> bytes:
        """Download file bytes from SharePoint."""
        logger.debug(f"Reading {file_path} from SharePoint")
        try:
            resp = requests.get(self._file_url(file_path), headers=self._get_headers(), timeout=30)
            resp.raise_for_status()
            logger.debug(f"Read {len(resp.content)} bytes from {file_path}")
            return resp.content
        except requests.RequestException as e:
            logger.error(f"Read failed for {file_path}: {e}")
            raise SharePointOperationError(f"Read failed for {file_path}: {e}")

    def write_excel(self, file_path: str, file_content: bytes) -> None:
        """Upload file bytes to SharePoint via a PUT."""
        logger.debug(f"Writing {len(file_content)} bytes to {file_path}")
        url = f"{self._drive_base()}/root:/{file_path}:/content"
        headers = self._get_headers()
        headers["Content-Type"] = "application/octet-stream"
        try:
            resp = requests.put(url, data=file_content, headers=headers, timeout=60)
            resp.raise_for_status()
            logger.debug(f"Wrote {file_path} successfully")
        except requests.RequestException as e:
            logger.error(f"Write failed for {file_path}: {e}")
            raise SharePointOperationError(f"Write failed for {file_path}: {e}")

    # ── Workbook API (incremental, co-authoring safe) ──

    def open_workbook_session(self, item_id: str, persist: bool = True) -> str:
        """Open a workbook session and return its id."""
        logger.info(f"Opening workbook session for item {item_id} (persist={persist})")
        url = f"{self._drive_base()}/items/{item_id}/workbook/createSession"
        try:
            resp = requests.post(url, json={"persistChanges": persist},
                                 headers=self._get_headers(), timeout=30)
            resp.raise_for_status()
            session_id = resp.json()["id"]
            logger.info(f"Workbook session opened: {_short_id(session_id)}")
            return session_id
        except requests.RequestException as e:
            logger.error(f"Failed to open workbook session for item {item_id}: {e}")
            raise SharePointOperationError(f"Failed to open workbook session: {e}")

    def close_workbook_session(self, item_id: str, session_id: str) -> None:
        logger.info(f"Closing workbook session {_short_id(session_id)}")
        url = f"{self._drive_base()}/items/{item_id}/workbook/closeSession"
        try:
            requests.post(url, headers=self._wb_headers(session_id), timeout=30)
            logger.info(f"Workbook session closed: {_short_id(session_id)}")
        except requests.RequestException as e:
            logger.warning(f"Failed to close workbook session {_short_id(session_id)}: {e}")

    def _wb_headers(self, session_id: str) -> dict:
        headers = self._get_headers()
        headers["workbook-session-id"] = session_id
        return headers

    def workbook_request(self, method: str, item_id: str, session_id: str,
                         rel_url: str, json_body: dict = None) -> dict:
        """Issue a Workbook API call relative to .../workbook/. Returns parsed JSON ({} if empty)."""
        url = f"{self._drive_base()}/items/{item_id}/workbook/{rel_url.lstrip('/')}"
        logger.debug(f"Workbook request: {method} {rel_url}")
        try:
            resp = requests.request(method, url, json=json_body,
                                    headers=self._wb_headers(session_id), timeout=60)
            resp.raise_for_status()
            return resp.json() if resp.content else {}
        except requests.HTTPError as e:
            body = ""
            try:
                body = e.response.json().get("error", {}).get("message", "")
            except Exception:
                pass
            logger.error(f"Workbook {method} {rel_url} failed: {e}" + (f" | {body}" if body else ""))
            raise SharePointOperationError(
                f"Workbook {method} {rel_url} failed: {e}" +
                (f" | {body}" if body else ""))
        except requests.RequestException as e:
            logger.error(f"Workbook {method} {rel_url} failed: {e}")
            raise SharePointOperationError(f"Workbook {method} {rel_url} failed: {e}")

    def archive_file(self, file_path: str, archive_folder: str = None) -> str:
        """Create a timestamped copy of a file on SharePoint.

        If a file with the same generated name (same sheet, same date) already
        exists anywhere under archive_folder -- including its subfolders -- its
        contents are overwritten in place instead of creating a duplicate.
        Otherwise the archive is written fresh at the root of archive_folder.
        """
        logger.info(f"Archiving {file_path}")
        try:
            file_bytes = self.read_excel(file_path)
            stem = Path(file_path).stem
            ext = Path(file_path).suffix
            date_suffix = datetime.now().strftime("%m%d%Y")
            archive_name = f"{stem}_{date_suffix}{ext}"

            if archive_folder:
                existing_path = self.find_file_path_in_tree(archive_folder, archive_name)
                dest = existing_path or f"{archive_folder}/{archive_name}"
            else:
                parent = str(Path(file_path).parent)
                dest = archive_name if parent == "." else f"{parent}/{archive_name}"

            self.write_excel(dest, file_bytes)
            logger.info(f"Archived {file_path} -> {dest}")
            return dest
        except SharePointOperationError:
            raise
        except Exception as e:
            logger.error(f"Archive failed for {file_path}: {e}")
            raise SharePointOperationError(f"Archive failed: {e}")

    # ── Convenience wrappers ──

    def load_workbook(self, file_path: str, keep_vba: bool = None):
        """Load an openpyxl Workbook from SharePoint."""
        if keep_vba is None:
            keep_vba = file_path.lower().endswith('.xlsm')
        wb_bytes = self.read_excel(file_path)
        return load_workbook(BytesIO(wb_bytes), keep_vba=keep_vba)

    def save_workbook(self, file_path: str, workbook) -> None:
        """Save an openpyxl Workbook back to SharePoint."""
        buf = BytesIO()
        workbook.save(buf)
        self.write_excel(file_path, buf.getvalue())

    def read_excel_to_dataframe(self, file_path: str):
        """Read SharePoint Excel file into a pandas DataFrame."""
        try:
            import pandas as pd
        except ImportError:
            raise SharePointOperationError("pandas not installed")
        content = self.read_excel(file_path)
        return pd.read_excel(BytesIO(content))

    def write_dataframe_to_excel(self, df, file_path: str) -> None:
        """Write a pandas DataFrame as Excel to SharePoint."""
        try:
            import pandas as pd
        except ImportError:
            raise SharePointOperationError("pandas not installed")
        buf = BytesIO()
        df.to_excel(buf, index=False, engine="openpyxl")
        self.write_excel(file_path, buf.getvalue())

